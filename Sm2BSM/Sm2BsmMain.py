from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment,Font,PatternFill
import glob
import tabula
import pandas as pd
import re


XLSX_files = glob.glob('*.xls*')
PDF_files = glob.glob('*.pdf')


import sys
import os
import openpyxl


if(len(sys.argv) < 2):
    print ('сканирую все файлы в каталоге')
    print (XLSX_files)
    print (PDF_files)
else:
    XLSX_files.clear()
    XLSX_files.append(sys.argv[1])
    print(sys.argv[0],'2')

NameShapka=["Поз","Поз","Наименование и техническая харрактеристика","Тип,марка,обозначение документа,опросного листа","Код продукции","Поставщик","Ед.измерения", "Кол.", "Масса 1 ед.,кг","Примечание"]
InpColumnASFS=[6,10,13,36,48,55, 64, 68,72,77]
InpColumnSEP=[ 7, 7, 11, 37, 49, 56, 65, 69, 73, 78]
#InpColumnPDF=[0,1,2,3,4,5, 6, 7,8,9]
InpColumnPDF=[1,2,3,4,5,6, 7, 8,8,10]
# InpColumnPDF=[3,4,5,6,7,8, 9, 10, 11, 12]


InpColumn=[6,10, 11, 12, 13, 20, 29, 30, 31, 36]
NumStolb = [1, 1, 2, 3, 4, 5, 6, 7, 8, 9]
PDFSTR = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '','']
Musor2 =['------','iiiiii'
        ]

Musor = ['nan', 'Позиция', 'Наименование и техническая характеристика', 'Тип, марка,\nобозначение документа,\nопросного листа                                ',
        'Код обору-\nдования, \nизделия,\nматериала', 'Поставщик', 'Еди-\nница\nизме-\nрения', 'Коли-\nчество', 'Масса \nединицы,\nкг', 'Примечание',
        'Поз.', 'Наименование и техническая характеристика',  '', 'Код\nпродукции', '', 'Ед.\nизме-\nрения', 'Кол.', 'Масса \n1 ед.,\nкг',
         'Поставщик', 'Ед.\rизме-\rрения', 'Изм.', '', 'Лист', 'Noдок.', 'Подп.', 'Дата','Рук.гр.', 'Рук.разд.','Н.контр.','ГИП',
         'одл.\rПоз.', 'одп. и датаВзам. инв. No\rНаименование и техническая характеристика', 'Тип, марка,\rобозначение документа,\rопросного листа', 'Код\rпродукции',
         'Подп. и датаВзам. инв. No\rНаименование и техническая характеристика', '', 'Код\rп  р  о  д  у  к  ц  и  и' ,'ол.уч', 'подл.\rПоз.'
         ,'Подп. Дата','Кол.уч','Проверил','Разраб.']
OutColumn=[1,2,3,4,5,6, 7, 8,9,10]
# PustStrok = [None, None, None, None, None, None, None, None, None, None]
PustStrok = ['', '', '', '', '', '', '', '', '', '']
PustStrok2 = ['-', '-', '-', '-', '-', '-', '-', '-', '-', '-']
OutRow = 10
OutString=[" "," "," "," "," "," "," ", " "," "," "]
NamWidth = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
ColWidth = [5, 11, 57, 32, 14, 26, 9, 10, 9, 25]

OutDir = "СПЕЦ"
Prefix = "\@_"

Prew=['Дорога:', 'Объект:' , 'Пусковой комплекс:','Номер документа:', 'Раздел проекта:','Проект:','Проекти-ровщик:','Другое:']


def PrintShapka(flag):
    # for j in range (1,8):
    #     OutCell = OutSheet.cell(row=j, column=1)
    #     OutCell.value = Prew[j]
    if flag=='XLS':
        cell = sheet.cell(row=50, column=61)
        OutCell = OutSheet.cell(row=4, column=3)
        OutCell.value = cell.value
        #
        cell = sheet.cell(row=52, column=61)
        OutCell = OutSheet.cell(row=6, column=3)
        OutCell.value = cell.value

        cell = sheet.cell(row=55, column=61)
        OutCell = OutSheet.cell(row=2, column=3)
        OutCell.value = cell.value
        cell = sheet.cell(row=58, column=61)
        OutCell = OutSheet.cell(row=5, column=3)
        OutCell.value = cell.value

    for j in range (1,10):
        OutCell = OutSheet.cell(row=OutRow, column=OutColumn[j])
        OutCell.value = NameShapka[j]
        OutCell.border = Border(top=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
        OutCell.alignment = Alignment(vertical='center',horizontal='center', wrap_text=True)

        OutCell.font = Font(name='TimesNewRoman', size = 12, bold = True)
        OutCell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type = "solid")
        OutCell = OutSheet.cell(row=OutRow+1, column=OutColumn[j])
        OutCell.value =  j
        OutCell.border = Border(top=Side(border_style='thin', color='000000'),
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))
        OutCell.alignment = Alignment(vertical='center',horizontal='center', wrap_text=True)

        OutCell.font = Font(name='TimesNewRoman', size = 12, bold = True)
        OutCell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type = "solid")
# устанавливаем ширину ячеек ( это надо перенестив шапку
    for i in range(0, 10):
        OutSheet.column_dimensions[NamWidth[i]].width = ColWidth[i]



##############################################################
#
#            Тестовая обработка одной таблицы
#
##############################################################



##############################################################
# Функция выделения номера из начала строки  вид X.XXX
#  косяк : может быть X.X.XXX
#  косяк : может быть X.XX.XXX
#  косяк : может быть XX.X.XXX
#  косяк : может быть XX.XX.XXX
##############################################################
def CutNumFromStringBegin( InputNum):
    OutputNum1=" "
    # return OutputNum1
    Num1 = InputNum[0:5]
    S = re.findall(r'\d+', Num1)
    #print (S)
    if (len(S) <= 0):
        OutputNum1 = ''
    elif len(S) == 1:
        OutString[1] = S[0]
    elif (Num1[1] != '.'):
        OutputNum1 = ''
    elif (len(S) == 1 and S[0] != 0):
        OutputNum1 = S[0]
    elif (len(S) == 2):
        OutputNum1 = S[0] + '.' + S[1]
    else:
        OutputNum1 = "?"
    return   OutputNum1

##############################################################
#  Функция определения в каком столбце название
#  Пример:
#  [0, 10, 3, 101, 66, 22, 13, 0, 0, 0]
#   0   1   2   3   4   5   6  7  8  9
#  вернет 3
##############################################################
def GetNumColSTring( ListOfCount):
#    ListOfCount2 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    maxEl = max(ListOfCount)
    return ListOfCount.index (maxEl)

##############################################################
#  Функция удаляет первый столбец если он пустой
##############################################################

def DelFirst0(nMYLIST):
    countpust=0
    n = len(nMYLIST)
    for i in range(0, n):
         if nMYLIST[i][0]=='':
             countpust+=1
    if(countpust==n):
         for i in range(0, n):
             nMYLIST[i].pop(0)

##############################################################
#  Функция добавляет первый столбец пустой
##############################################################

def InsFirst0(nMYLIST):
    n = len(nMYLIST)
    for i in range(0, n):
         nMYLIST[i].insert(0,"")


def ParsingOnePDF():
    MYLIST = [["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
              ["", "", "", "", "", "", "", "", "", "", "", "", "", ""],
              ["", "", "", "", "", "", "", "", "", "", "", "", "", ""]]
    OutRows = 1
    MaxStr =7
    OutStringP = ["", "", "", "", "", "", "", "", "", ""]
#    for i in range(500):
        #MYLIST.append(range(0, 12))
#        MYLIST.append(["", "", "", "", "", "", "", "", "", ""])
#    for i in range(0, 500):
#        print(MYLIST[i])
    print (rows, len (PDFT),len (PDFT.iloc[0]))
    for i in range(0, len (PDFT) -1):
        # Здесь нужно внимательно посмотреть не до 7 а до скольки ?
        for j in range(0, MaxStr):
            # print ("индексы [", i,j ,']')
            OutStringP[j] = str(PDFT.iloc[i,  j])
            FlagMusor=0
#  чистим мусор по контенту
            for L in range( 0,len(Musor)):
                if (OutStringP[j] == Musor[L] ):
                    OutStringP[j] = ''
                    FlagMusor=1
# если в строке был мусор - чистим строку совсем
            if(FlagMusor==1):
                for L in range( 0,MaxStr):
                    OutStringP[j] = ''

        if  OutStringP[0]!='':
            for L in range(0, len(OutStringP)):
                OutStringP[L] = ''

        if OutStringP.count('')!= len(OutStringP):
            MYLIST.append( ["", "", "", "", "", "", "", "", "", "", "", "", "", ""])
            OutRows += 1
            for j in range(0, MaxStr):
                 MYLIST[OutRows-1][j] = str(OutStringP[j])
# нормируем
    DelFirst0(MYLIST)
    DelFirst0(MYLIST)
    DelFirst0(MYLIST)
    InsFirst0(MYLIST)
    InsFirst0(MYLIST)
#   определяем в каком столбце больше всего записей -  это столбец 2 "наименование"
    CountStrok = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    for i in range(0, OutRows):
        for j in range(0, MaxStr):
            if MYLIST[i][j] !='':
                CountStrok[j]+=1
    print(CountStrok,GetNumColSTring(CountStrok))
#  Правильный вариант
    if GetNumColSTring(CountStrok)==0 :
        InsFirst0(MYLIST)
        InsFirst0(MYLIST)

    if GetNumColSTring(CountStrok)==1 :
        InsFirst0(MYLIST)

    if (GetNumColSTring(CountStrok) == 3):
        DelFirst0(MYLIST)
    if (GetNumColSTring(CountStrok) == 4):
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)
    if (GetNumColSTring(CountStrok) == 5):
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)
    if (GetNumColSTring(CountStrok) == 6):
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)
        DelFirst0(MYLIST)

    PrintShapka('PDF')
    #    сюда помещу вырезание квадрата



    #   стираем строку если толко первый 6й заполнен
    for i in range(0, OutRows):
        count_NNUL = 0
        index_NNUL = 0
        for j in range(0, MaxStr):
            if MYLIST[i][j] !='':
                count_NNUL+=1
                index_NNUL=j
        if count_NNUL==1 and index_NNUL==6 :
            MYLIST[i][6] = ''
        if count_NNUL == 1 and index_NNUL == 5:
            MYLIST[i][5] = ''
        if count_NNUL == 1 and index_NNUL == 7:
            MYLIST[i][7] = ''
        if count_NNUL == 1 and index_NNUL == 8:
            MYLIST[i][8] = ''

    #    выделяем номер в первый столбец
    for i in range(0, OutRows):
         MYLIST[i][1] = CutNumFromStringBegin(MYLIST[i][2])
         if(MYLIST[i][1]!=MYLIST[i][2]):
            MYLIST[i][2] = MYLIST[i][2].replace(MYLIST[i][1],'')
         else:
             MYLIST[i][1] !=''
    #   объеденяем строчки и пытаемся вытащить то что улетело )))
    for i in range(0, len (MYLIST)-6):
      #
        if(MYLIST[i][1]!=''):
            for k in range(i+1,i+6):

                if (MYLIST[k][1] == ''):
                    # Сначала следует проверить ???  а нет ли справа числа нормального ?
                    if (CutNumFromStringBegin(MYLIST[k][3])!= '') and (CutNumFromStringBegin(MYLIST[k][3])!= ' ') :
                        print (CutNumFromStringBegin(MYLIST[k][3]))
                        MYLIST[k].pop(0)
                        MYLIST[k][1] = CutNumFromStringBegin(MYLIST[k][2])
                        MYLIST[i][2] = MYLIST[i][2].replace(MYLIST[i][1], '')
                    else:
                        MYLIST[i][2] = MYLIST[i][2] + "\n"+MYLIST[k][2]
                        MYLIST[k][2]=''
                        MYLIST[i][3]+= "\n"+MYLIST[k][3]
                        MYLIST[k][3]=''
                        MYLIST[i][4]+="\n"+MYLIST[k][4]
                        MYLIST[k][4]=''
                        MYLIST[i][5]+="\n"+MYLIST[k][5]
                        MYLIST[k][5]=''
                        MYLIST[i][6]+="\n"+MYLIST[k][6]
                        MYLIST[k][6]=''
                else:
                    break
    #  Убираем пустые строки
    #   стираем строку если толко первый 6й заполнен
    del_index = 0
    for i in range(0, len(MYLIST)):
        count_NNUL = 0
        for j in range(0, MaxStr):
            if MYLIST[del_index][j] !='':
               count_NNUL+=1
        if count_NNUL==0 :
            MYLIST.pop(del_index)
        else:
            del_index+=1
    #    Переписываем наш массив в выходной файл экселя
    for i in range(0, len(MYLIST)):
        print(MYLIST[i])
        for j in range(0, MaxStr):
            OutCell = OutSheet.cell(row=i+12, column=j+1)
            OutCell.value = MYLIST[i ][j]
            #  Границы, Шрифт, расположение
            OutCell.border = Border(top=Side(border_style='thin', color='000000'), \
                                    left=Side(border_style='thin', color='000000'), \
                                    right=Side(border_style='thin', color='000000'), \
                                    bottom=Side(border_style='thin', color='000000')
                                    )
            if j==2 :
                OutCell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
            else:
                OutCell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            OutCell.font = Font(name='TimesNewRoman', size=10, bold=False)



    return OutRows


def ParsingOneFile(flag):
    OutRows = OutRow+1
    FlagRzadelNULLLine = 1

    for i in range(1, rows + 1):

    # это для  PDF
        if flag=='PDF' :
            for j in range(0, 7):
                OutString[j] = str(pdf_tables[0].iloc[i,  InpColumn[j]])
                for L in range( 0,len(Musor)):
                    if (OutString[j] == Musor[L] ):
                        OutString[j] = ''

            #    print(OutString)
            Num1 = OutString[2][0:5]

            S = re.findall(r'\d+', Num1)
            if len(S) <= 0 :
                OutString[1] = ''
            elif len(S) == 1:
                OutString[1] = S[0]
            elif (Num1[1] != '.'):
                OutString[1] = ''
            elif (len(S) == 1 and S[0] != 0):
                OutString[1] = S[0]
            elif (len(S) == 2):

                OutString[1] = S[0] + '.' + S[1]
            else:
                OutString[1] = "11"

            # print(OutString)
# это для  XLS
        if flag=='XLS' :
            for j in range (0,10):
                OutString[j] = ''
                cell = sheet.cell (row = i, column = InpColumn[j])
                OutString[j] = cell.value
                # OutCell = OutSheet.cell(row=OutRows, column=OutColumn[i])
                # OutCell.value =  str("")

                for L in range( 0,len(Musor)):
                    if (OutString[j] == Musor[L] ):
                        OutString[j] = ''

            print(OutString)

        if(OutString==NumStolb):
            for  N in range(0,len(OutString)):
                OutString[N] = ""
        if (FlagRzadelNULLLine == 1) and (OutString[2] != None) and (OutString[2] != '')  and (OutString[1] == ''):
        # Это начало раздела ,
            OutRows+=1

            FlagRzadelNULLLine == 0
            OutCell = OutSheet.cell(row=OutRows, column=OutColumn[1])
            OutCell.border = Border(  left=Side(border_style='thin', color='000000'))
            OutCell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

            OutCell = OutSheet.cell(row=OutRows, column=OutColumn[9])
            OutCell.border = Border(right=Side(border_style='thin', color='000000'))
            OutCell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
            OutCell = OutSheet.cell(row=OutRows, column=OutColumn[2])
            OutCell.font = Font(name='TimesNewRoman', size=12, bold=True)
            OutCell.value = str(OutString[2])

        if  (OutString[1]!=None) and (OutString[1]!="")  and (OutString[2]!=None) and (OutString[2]!=""):
            OutRows+=1
            for j in range(1, 10):
                OutCell = OutSheet.cell (row = OutRows, column = OutColumn[j])
                OutCell.value = OutString[j]
                OutCell.alignment = Alignment(vertical='center',horizontal='left', wrap_text=True)
                OutCell.font = Font(name='TimesNewRoman', size = 10, bold = False)
                OutCell.border = Border( top=Side(border_style='thin', color='000000'),\
                                         left = Side(border_style='thin', color='000000'),\
                                         right = Side(border_style='thin', color='000000'), \
                                         bottom = Side(border_style='thin', color='000000')
                                         )
                if(j!=2) and (j!=3) :
                    OutCell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                else:
                    OutCell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

        if (FlagRzadelNULLLine == 0)  and (OutString[2]!=None) and  (OutString[2]!="")  and ((OutString[1]==None) or (OutString[1]=='')):
            OutCell = OutSheet.cell (row = OutRows, column = OutColumn[2])

            if  (OutCell.value!=2 ) :
                OutCell.value = str(OutCell.value) + str( OutString[2])
            else:
                # херовво я тут сделал  , просто скопировал кусок
                OutRows += 1

                FlagRzadelNULLLine == 0
                OutCell = OutSheet.cell(row=OutRows, column=OutColumn[1])
                OutCell.border = Border(left=Side(border_style='thin', color='000000'))
                OutCell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

                OutCell = OutSheet.cell(row=OutRows, column=OutColumn[9])
                OutCell.border = Border(right=Side(border_style='thin', color='000000'))
                OutCell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                OutCell = OutSheet.cell(row=OutRows, column=OutColumn[2])
                OutCell.font = Font(name='TimesNewRoman', size=12, bold=True)
                OutCell.value = str(OutString[2])

        # if (OutString[3] != None) and (OutString[3] != "") and (FlagRzadel==1)and ((OutString[1]==None) or  (OutString[1]=="")) :
        if (OutString[3] != None) and (OutString[3] != "") and ((OutString[1]==None) or  (OutString[1]=="")) :
            OutCell = OutSheet.cell(row=OutRows, column=OutColumn[3])
            OutCell.value = str(OutCell.value) +"\n" + str(OutString[3])




        if ( (OutString[1]=='')and (OutString[2]=='')and(OutString[3]=='')and(OutString[4]=='')and(OutString[5]=='')and(OutString[6]=='')) :
             FlagRzadelNULLLine = 1
        else :
            FlagRzadelNULLLine = 0

#    Установить ширину ячеек
    for i in range(0, 10):
        OutSheet.column_dimensions[NamWidth[i]].width = ColWidth[i]
    OutRows += 1
    return OutRows



for InputFilename in  XLSX_files :
# выходной файл
    print('Обрабатываем ' + InputFilename)
    OutputFilename = OutDir + Prefix + InputFilename
    OutputFilename = OutputFilename.replace("xlsm", "xlsx")
    if ".xlsm" in InputFilename:
        InpColumn = InpColumnSEP
        # for  N in range(0,len(InpColumn)):
        #     InpColumn[N] = InpColumnSEP[N]
    else:
        InpColumn = InpColumnASFS
        # for N in range(0, len(InpColumn)):
        #     InpColumn[N] = InpColumnASFS[N]

#    print(OutputFilename)
# читаем excel-файл
    try:
        wb1 = openpyxl.load_workbook(InputFilename)
        sheet = wb1.active
        rows = sheet.max_row
        cols = sheet.max_column
        wbOut = Workbook()

        # grab the active worksheet
        OutSheet = wbOut.active

        # печатаем шапку
        PrintShapka('XLS')
        # парсим
        OR = ParsingOneFile('XLS')
        # записываем       print('Сохраняем ' + OutputFilename)

        try:
            os.mkdir(OutDir)
        except FileExistsError:
            print('каталог уже существует')
        OutCell = OutSheet.cell(row=OR, column=2)
        OutCell.value = "Обязательно сверьте полученный результат с исходным докуметом.  (c)GTSS @MAA  62115"

        # Сохранение выходного файла
        try:
            wbOut.save(OutputFilename)
        except:
            print('немогу записать')


    except FileExistsError:
    # except :
        print ("Проблема с файлом ", InputFilename, "не обработан")

# теперь повторяем все тожесамое но для pdf
NY=0
for InputFilename in  PDF_files :
# выходной файл
    print('Обрабатываем ' + InputFilename)
    OutputFilename = OutDir + Prefix + InputFilename
    OutputFilename = OutputFilename.replace("pdf", "xlsx")
    InpColumn = InpColumnPDF

# читаем pdf-файл
    try:

        # читаем pdf
        pdf_tables = tabula.read_pdf(InputFilename,
                                     #                  Y(30)   X   Y1  X1
                                     pages='all', area=[30,50,720,1000],
                                      multiple_tables=True)
        #  multiple_tables = False)
        ind=0
        for z    in pdf_tables:
            ind +=1
            z.to_excel(str(ind) + 'X.xlsx')
        # Convert into Excel File
        # pdf_tables[0].to_excel('1.xlsx')
        print ('ВСЕГО ТАБЛИЦ: ', len (pdf_tables))
        NTab = 0

        for PDFT in pdf_tables:
            NTab+=1
            print ("таблица=",NTab,"СТРОК=",len(PDFT),"СТОЛБЦОВ=",len(PDFT.iloc[0]))

            for o in range (0,len(PDFT)):
                for j in range(0, len(PDFT.iloc[o])):
                    PDFSTR[j] = str(PDFT.iloc[o, j])
                print (PDFSTR)

        wbOut = Workbook()
        # grab the active worksheet
        OutSheet = wbOut.active

        # печатаем шапку
###########        PrintShapka('PDF')
        # парсим
        if(len(pdf_tables)>0):
            rows = len (pdf_tables[0])-1
            PDFT=pdf_tables[0]
            # PDFT = PDFT.append( ['--', '--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--','--',])
            for z in pdf_tables:
                PDFT=z
            #     PDFT = PDFT.append(z)

            #        print (InpColumn)
#######     OR = ParsingOneFile('PDF')
#                OR = ParsingOnePDF()
        # записываем
            print('Сохраняем ' + OutputFilename)

            try:
                os.mkdir(OutDir)
            except FileExistsError:
                print('каталог уже существует')
            OutCell = OutSheet.cell(row=OR, column=2)
        # OutCell.value = "Обязательно сверьте полученный результат с исходным докуметом.  (c)GTSS @MAA  62115"

        # Сохранение выходного файла
            try:
                NY+=1
                wbOut.save(str(NY)+OutputFilename)
            except:
                print('немогу записать')


    except FileExistsError:
    #except :
        print ("Проблема с файлом ", InputFilename, "не обработан")





    # print(OutString[1],OutString[2],OutString[3],OutString[4],OutString[5])
print('Работа завершена')
